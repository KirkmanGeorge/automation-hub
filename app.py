import os, time, re, glob
import streamlit as st
import openpyxl
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import random
from io import BytesIO
import difflib

st.set_page_config(page_title="Automation Hub", layout="wide", page_icon="🤖")

st.markdown("""
<style>
    .stApp { background-color: #F3F3F3; color: #000000; font-family: 'Segoe UI', sans-serif; }
    h1, h2, h3 { color: #0078D4; font-weight: 600; }
    .stButton > button {
        background-color: #0078D4; color: white; border-radius: 4px;
        border: none; padding: 8px 16px; font-weight: 500;
    }
    .stButton > button:hover { background-color: #106EBE; }
    p, li, span, div { color: #000000 !important; }
    .log-box {
        background: #1e1e1e; color: #00ff88; font-family: monospace;
        font-size: 13px; padding: 12px 16px; border-radius: 6px;
        max-height: 400px; overflow-y: auto; white-space: pre-wrap;
    }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# TOOL 1: Excel Stock Movement Filler
# ─────────────────────────────────────────────────────────────────────────────

def normalize_name(name):
    if not name:
        return ""
    return str(name).upper().replace(" ", "").replace("S", "")

def excel_serial_to_date(serial):
    if isinstance(serial, (float, int)):
        return (datetime(1899, 12, 30) + timedelta(days=serial)).date()
    if isinstance(serial, datetime):
        return serial.date()
    return None

def process_excel(template_file, report_file, damages_file, output_name="filled_template.xlsx"):
    try:
        template_bytes = template_file.read()
        wb_temp = openpyxl.load_workbook(BytesIO(template_bytes))
        ws_temp = wb_temp['Sheet1']
        template_first_date = None
        for r in range(1, ws_temp.max_row + 1):
            date_val = ws_temp.cell(r, 1).value
            if date_val:
                template_first_date = excel_serial_to_date(date_val)
                if template_first_date:
                    break
        if not template_first_date:
            raise ValueError("No dates found in template")
        template_year = template_first_date.year
        report_df  = pd.read_excel(report_file)
        damages_df = pd.read_excel(damages_file)
        report_df['date'] = pd.to_datetime(report_df['date'], dayfirst=True)
        report_df['date'] = report_df['date'].apply(lambda d: d.replace(year=template_year))
        report_df_sorted = report_df.sort_values(['abbreviations', 'date'])
        adj_df = report_df_sorted[report_df_sorted['movement_type'] == 'Stock adjustment']
        for _, adj_row in adj_df.iterrows():
            abr      = adj_row['abbreviations']
            adj_date = adj_row['date']
            adj_amt  = abs(adj_row['adjusted amount'])
            same_day_ins = report_df_sorted[
                (report_df_sorted['abbreviations'] == abr) &
                (report_df_sorted['date'] == adj_date) &
                (report_df_sorted['movement_type'] == 'Stock-in')]
            if not same_day_ins.empty:
                i = same_day_ins.index[-1]
                report_df_sorted.at[i, 'adjusted amount'] = max(0, report_df_sorted.at[i, 'adjusted amount'] - adj_amt)
                continue
            prev_ins = report_df_sorted[
                (report_df_sorted['abbreviations'] == abr) &
                (report_df_sorted['date'] < adj_date) &
                (report_df_sorted['movement_type'] == 'Stock-in')]
            if not prev_ins.empty:
                i = prev_ins.index[-1]
                report_df_sorted.at[i, 'adjusted amount'] = max(0, report_df_sorted.at[i, 'adjusted amount'] - adj_amt)
        ins_df  = report_df_sorted[report_df_sorted['movement_type'] == 'Stock-in'].groupby(
            ['date', 'abbreviations'])['adjusted amount'].sum().reset_index(name='stock_in')
        outs_df = report_df_sorted[report_df_sorted['movement_type'] == 'Invoice Issue'].groupby(
            ['date', 'abbreviations'])['adjusted amount'].sum().reset_index(name='sales')
        first_appearances = report_df_sorted.drop_duplicates(subset=['abbreviations'], keep='first')
        openings = dict(zip(first_appearances['abbreviations'], first_appearances['book quantity']))
        damages_df = damages_df[pd.notna(damages_df['quantity'])]
        wb = openpyxl.load_workbook(BytesIO(template_bytes))
        ws = wb['Sheet1']
        product_map = {}
        norm_to_abr = {}
        for r in range(1, ws.max_row + 1):
            full = ws.cell(r, 2).value
            abr  = ws.cell(r, 3).value
            if full and abr:
                product_map[str(full).strip().upper()] = abr
                norm_to_abr[normalize_name(full)] = abr
        damages_dict = {}
        for _, drow in damages_df.iterrows():
            full = str(drow['good name']).strip().upper()
            qty  = int(drow['quantity'])
            if full in product_map:
                damages_dict[product_map[full]] = qty
            else:
                norm_full = normalize_name(full)
                if norm_full in norm_to_abr:
                    damages_dict[norm_to_abr[norm_full]] = qty
        report_abr_map = {}
        for _, rrow in report_df.iterrows():
            full       = str(rrow['good name']).strip().upper()
            abr_report = rrow['abbreviations']
            if full in product_map:
                report_abr_map[abr_report] = product_map[full]
            else:
                norm_full = normalize_name(full)
                if norm_full in norm_to_abr:
                    report_abr_map[abr_report] = norm_to_abr[norm_full]
        ins_df['abbreviations']  = ins_df['abbreviations'].map(report_abr_map).fillna(ins_df['abbreviations'])
        outs_df['abbreviations'] = outs_df['abbreviations'].map(report_abr_map).fillna(outs_df['abbreviations'])
        mapped_openings = {report_abr_map.get(k, k): v for k, v in openings.items()}
        date_abr_to_row = {}
        current_date = None
        for r in range(1, ws.max_row + 1):
            date_val = ws.cell(r, 1).value
            if date_val:
                current_date = excel_serial_to_date(date_val)
            abr = ws.cell(r, 3).value
            if current_date and abr:
                date_abr_to_row[(current_date, abr)] = r
        if date_abr_to_row:
            first_date = min(d[0] for d in date_abr_to_row)
            for abr, open_bal in mapped_openings.items():
                key = (first_date, abr)
                if key in date_abr_to_row:
                    ws.cell(date_abr_to_row[key], 4).value = open_bal
        damages_per_day = {}
        for abr, total_d in damages_dict.items():
            abr_ins = ins_df[ins_df['abbreviations'] == abr]
            if abr_ins.empty:
                continue
            days           = abr_ins['date'].dt.date.values
            stock_ins      = abr_ins['stock_in'].values
            total_stock_in = stock_ins.sum()
            if total_stock_in == 0:
                continue
            weights    = stock_ins / total_stock_in
            prod_d     = int(total_d * 3 / 4)
            pack_d     = total_d - prod_d
            prod_alloc = np.zeros(len(days), dtype=int)
            pack_alloc = np.zeros(len(days), dtype=int)
            for _ in range(prod_d):
                prod_alloc[np.random.choice(len(days), p=weights)] += 1
            for _ in range(pack_d):
                pack_alloc[np.random.choice(len(days), p=weights)] += 1
            damages_per_day[abr] = {days[i]: (prod_alloc[i], pack_alloc[i]) for i in range(len(days))}
        for _, irow in ins_df.iterrows():
            dt       = irow['date'].date()
            abr      = irow['abbreviations']
            stock_in = irow['stock_in']
            key      = (dt, abr)
            if key not in date_abr_to_row:
                continue
            row_num = date_abr_to_row[key]
            prod_d_day, pack_d_day = damages_per_day.get(abr, {}).get(dt, (0, 0))
            total_d_day = prod_d_day + pack_d_day
            ws.cell(row_num, 7).value  = stock_in + total_d_day
            ws.cell(row_num, 8).value  = prod_d_day
            ws.cell(row_num, 10).value = pack_d_day
            actual_filled = stock_in + total_d_day
            if actual_filled > 0:
                if actual_filled <= 50:    diff = random.randint(-1, 1)
                elif actual_filled <= 200: diff = random.randint(-4, 6)
                else:                      diff = random.randint(-6, 12)
                ws.cell(row_num, 6).value = max(0, actual_filled + diff)
        for _, orow in outs_df.iterrows():
            dt  = orow['date'].date()
            abr = orow['abbreviations']
            key = (dt, abr)
            if key in date_abr_to_row:
                ws.cell(date_abr_to_row[key], 13).value = orow['sales']
        output_bytes = BytesIO()
        wb.save(output_bytes)
        output_bytes.seek(0)
        return output_bytes
    except Exception as e:
        st.error(f"Error: {str(e)}")
        return None


# ─────────────────────────────────────────────────────────────────────────────
# TOOL 2: EFRIS Invoice Enricher
# ─────────────────────────────────────────────────────────────────────────────

def _parse_pdf_bytes(pdf_bytes):
    import pdfplumber, io
    items = []
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                for table in (page.extract_tables() or []):
                    header_idx, col_map = None, {}
                    for i, row in enumerate(table):
                        if not row:
                            continue
                        upper = [str(c or "").upper().strip() for c in row]
                        if "ITEM" in upper and "QUANTITY" in upper:
                            header_idx = i
                            for j, h in enumerate(upper):
                                if h == "ITEM":                       col_map["item"]         = j
                                elif h == "QUANTITY":                 col_map["quantity"]     = j
                                elif "UNIT" in h and "MEASURE" in h: col_map["unit_measure"] = j
                                elif "UNIT" in h and "PRICE" in h:   col_map["unit_price"]   = j
                            continue
                        if header_idx is not None:
                            def g(key, fb, row=row):
                                ix = col_map.get(key, fb)
                                return str(row[ix] or "").strip() if ix < len(row) else ""
                            item_name = g("item", 1)
                            qty = g("quantity", 2)
                            if item_name and qty and not item_name.upper().startswith("TAX"):
                                items.append({
                                    "item": item_name,
                                    "quantity": qty,
                                    "unit_measure": g("unit_measure", 3),
                                    "unit_price": g("unit_price", 4),
                                })
                    if items:
                        return items
                text = page.extract_text() or ""
                in_d = False
                for line in text.split("\n"):
                    line = line.strip()
                    if "Section D" in line or "Goods & Services" in line:
                        in_d = True
                        continue
                    if "Section E" in line or "Tax Details" in line:
                        break
                    if not in_d:
                        continue
                    m = re.match(
                        r"\d+\.?\s+(.+?)\s+(\d[\d,]*)\s+(\S[\S\-]*)\s+([\d,]+)\s+[\d,]+",
                        line
                    )
                    if m:
                        items.append({
                            "item": m.group(1).strip(),
                            "quantity": m.group(2).strip(),
                            "unit_measure": m.group(3).strip(),
                            "unit_price": m.group(4).strip(),
                        })
    except Exception as e:
        print(f"[PDF PARSE ERROR] {e}")
    return items


def _get_driver():
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-setuid-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-gpu-sandbox")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--no-zygote")
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-software-rasterizer")
    options.add_argument("--disable-features=VizDisplayCompositor")
    options.add_argument("--disable-background-timer-throttling")
    options.add_argument("--disable-backgrounding-occluded-windows")
    options.add_argument("--disable-renderer-backgrounding")
    options.set_capability("goog:loggingPrefs", {"performance": "ALL"})
    browser_candidates = [
        "/usr/bin/google-chrome", "/usr/bin/google-chrome-stable",
        "/usr/lib/chromium/chromium", "/usr/bin/chromium", "/usr/bin/chromium-browser",
    ]
    driver_candidates = [
        "/usr/bin/chromedriver", "/usr/local/bin/chromedriver", "/usr/lib/chromium/chromedriver",
    ]
    browser   = next((p for p in browser_candidates if os.path.exists(p)), None)
    driver_bin = next((p for p in driver_candidates if os.path.exists(p)), None)
    if not browser:
        hits = glob.glob("/usr/**/chrome", recursive=True) + glob.glob("/usr/**/chromium", recursive=True)
        browser = next((h for h in hits if os.access(h, os.X_OK)), None)
    if not driver_bin:
        hits = glob.glob("/usr/**/chromedriver", recursive=True)
        driver_bin = next((h for h in hits if os.access(h, os.X_OK)), None)
    if browser:
        options.binary_location = browser
    if driver_bin:
        return webdriver.Chrome(service=Service(executable_path=driver_bin), options=options)
    return webdriver.Chrome(options=options)


def _scrape_fdn(driver, fdn, log_fn=None):
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    import requests as req_lib, json

    def dbg(msg):
        print(msg)
        if log_fn:
            log_fn(msg)

    items = []
    original_handle = driver.current_window_handle
    try:
        driver.get("https://efris.ura.go.ug/")
        wait = WebDriverWait(driver, 20)
        dbg("  [1] Loaded EFRIS")
        inp = wait.until(EC.presence_of_element_located((By.XPATH,
            "//input[@placeholder and ("
            "contains(translate(@placeholder,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'fiscal')"
            " or contains(translate(@placeholder,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'fdn')"
            ")]"
        )))
        inp.clear(); inp.send_keys(str(fdn)); time.sleep(0.4)
        dbg("  [2] FDN typed")
        btn = wait.until(EC.element_to_be_clickable((By.XPATH,
            "//button[contains(translate(normalize-space(.),"
            "'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'VALIDATE')]"
        )))
        btn.click(); time.sleep(3); dbg("  [3] Validated")
        wait.until(EC.presence_of_element_located((By.XPATH,
            "//*[contains(text(),'erified') or contains(text(),'Validation')]"
        ))); time.sleep(1); dbg("  [4] Invoice verified")
        handles_before = set(driver.window_handles)
        vbtn = wait.until(EC.element_to_be_clickable((By.XPATH,
            "//button[contains(translate(normalize-space(.),"
            "'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'VIEW DOCUMENT')]"
        )))
        vbtn.click(); time.sleep(5); dbg("  [5] View Document clicked")
        pdf_url = None
        try:
            logs = driver.get_log("performance")
            for entry in logs:
                msg = json.loads(entry["message"])["message"]
                if msg.get("method") == "Network.responseReceived":
                    url  = msg.get("params", {}).get("response", {}).get("url", "")
                    mime = msg.get("params", {}).get("response", {}).get("mimeType", "")
                    if "pdf" in mime.lower() or (url and ".pdf" in url.lower()):
                        pdf_url = url; break
        except Exception as e:
            dbg(f"  [6-CDP] Log error: {e}")
        if not pdf_url:
            new_handles = set(driver.window_handles) - handles_before
            if new_handles:
                driver.switch_to.window(list(new_handles)[0]); time.sleep(2)
                pdf_url = driver.current_url
                driver.close(); driver.switch_to.window(original_handle)
        if not pdf_url:
            src = driver.page_source
            for u in re.findall(r'https?://[^\s"\'<>\\]+', src):
                if ".pdf" in u.lower() or "printInvoice" in u or "viewDoc" in u:
                    pdf_url = u; break
        if not pdf_url:
            for tag in ["embed", "iframe", "object", "a"]:
                for el in driver.find_elements(By.TAG_NAME, tag):
                    for attr in ["src", "data", "href"]:
                        val = el.get_attribute(attr) or ""
                        if val and ("pdf" in val.lower() or "invoice" in val.lower()):
                            pdf_url = val; break
                    if pdf_url: break
                if pdf_url: break
        pdf_bytes = None
        if pdf_url and pdf_url.startswith("http"):
            cookies = {c["name"]: c["value"] for c in driver.get_cookies()}
            headers = {"User-Agent": "Mozilla/5.0", "Referer": "https://efris.ura.go.ug/"}
            try:
                resp = req_lib.get(pdf_url, cookies=cookies, headers=headers, timeout=30)
                if resp.status_code == 200 and len(resp.content) > 500:
                    pdf_bytes = resp.content
            except Exception as e:
                dbg(f"  [8] Download error: {e}")
        elif pdf_url and pdf_url.startswith("blob:"):
            try:
                js  = ("var cb=arguments[arguments.length-1];"
                       "fetch(arguments[0]).then(r=>r.arrayBuffer())"
                       ".then(b=>cb(Array.from(new Uint8Array(b)))).catch(e=>cb([]));")
                arr = driver.execute_async_script(js, pdf_url)
                if arr:
                    pdf_bytes = bytes(arr)
            except Exception as e:
                dbg(f"  [8] Blob error: {e}")
        if pdf_bytes:
            items = _parse_pdf_bytes(pdf_bytes)
    except Exception as e:
        dbg(f"  [ERR] {e}")
    try:
        if driver.current_window_handle != original_handle:
            driver.switch_to.window(original_handle)
    except Exception:
        pass
    return items


def fuzzy_match(target, candidates):
    t  = target.strip().upper()
    cs = [c.strip().upper() for c in candidates]
    ms = difflib.get_close_matches(t, cs, n=1, cutoff=0.55)
    return candidates[cs.index(ms[0])] if ms else None


def run_efris_enrichment(purchases_df, log_placeholder, progress_bar):
    for col in ["Quantity", "Unit Measure", "Unit Price"]:
        if col not in purchases_df.columns:
            purchases_df[col] = None
    total, log_lines = len(purchases_df), []
    def log(msg):
        log_lines.append(msg)
        log_placeholder.markdown(
            '<div class="log-box">' + "<br>".join(log_lines[-80:]) + "</div>",
            unsafe_allow_html=True)
    import shutil
    for name in ["chromium", "chromium-browser", "chromedriver"]:
        p = shutil.which(name) or next(
            (c for c in [f"/usr/lib/chromium/{name}", f"/usr/bin/{name}"] if os.path.exists(c)), None)
        log(f"{'✅' if p else '❌'}  {name} → {p or 'not found'}")
    log("🚀 Starting browser...")
    try:
        driver = _get_driver()
        log("✅ Browser started!")
    except Exception as e:
        st.error(f"Browser failed: {e}")
        return purchases_df
    fdn_cache = {}
    try:
        for idx, row in purchases_df.iterrows():
            fdn  = str(row.get("FDN", "")).strip()
            desc = str(row.get("Description of Goods", "")).strip()
            row_num = idx + 2
            progress_bar.progress(min((idx + 1) / total, 1.0), text=f"Row {idx+1}/{total} — {fdn}")
            if not fdn or fdn.lower() == "nan":
                log(f"[Row {row_num}] ⚠️  Skipped — no FDN"); continue
            if fdn not in fdn_cache:
                log(f"[Row {row_num}] 🔍  FDN: {fdn} | {desc}")
                try:
                    fdn_cache[fdn] = _scrape_fdn(driver, fdn, log_fn=log)
                    log(f"[Row {row_num}] ✅  {len(fdn_cache[fdn])} item(s) found")
                except Exception as e:
                    fdn_cache[fdn] = []; log(f"[Row {row_num}] ❌  {e}")
            else:
                log(f"[Row {row_num}] 📋  Cached — {fdn}")
            invoice_items = fdn_cache[fdn]
            if not invoice_items:
                continue
            invoice_names = [i["item"] for i in invoice_items]
            matched = fuzzy_match(desc, invoice_names)
            if matched:
                hit = next((i for i in invoice_items if i["item"].strip().upper() == matched.strip().upper()), None)
                if hit:
                    purchases_df.at[idx, "Quantity"]     = hit["quantity"]
                    purchases_df.at[idx, "Unit Measure"] = hit["unit_measure"]
                    purchases_df.at[idx, "Unit Price"]   = hit["unit_price"]
                    log(f"[Row {row_num}] ✔️  '{desc}' → Qty:{hit['quantity']} Unit:{hit['unit_measure']} Price:{hit['unit_price']}")
                else:
                    log(f"[Row {row_num}] ⚠️  Lookup failed: {matched}")
            else:
                log(f"[Row {row_num}] ⚠️  No match for '{desc}'")
    finally:
        try:
            driver.quit()
        except Exception:
            pass
    log("🏁 All rows processed.")
    return purchases_df


def build_output_excel(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Purchases Report")
        ws = writer.sheets["Purchases Report"]
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 50)
        from openpyxl.styles import PatternFill, Font, Alignment
        hdr_fill = PatternFill("solid", fgColor="0078D4")
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        hi_fill  = PatternFill("solid", fgColor="FFF2CC")
        new_cols = {"Quantity", "Unit Measure", "Unit Price"}
        for i, c in enumerate(ws[1]):
            if c.value in new_cols:
                for row in ws.iter_rows(min_row=2, min_col=i+1, max_col=i+1):
                    for cell in row:
                        cell.fill = hi_fill
    output.seek(0)
    return output


# ─────────────────────────────────────────────────────────────────────────────
# TOOL 3: Raw Material Movement Filler
# ─────────────────────────────────────────────────────────────────────────────

# Header / output row keywords that are never raw material or product names
_SM_HEADER_WORDS = {
    'details', 'units', 'output', 'input', 'qty', 'price', 'amount',
    'unit', 'measure', 'total', 'description', 'items', 'sub',
    'out put', 'out', 'put',
}

def _norm_name(s):
    """Uppercase + strip — used as the matching key for product names."""
    return str(s).strip().upper()


def _detect_sm_structure(df):
    """
    Dynamically locate the three key rows in ANY company's standard mix:
      product_name_row   — first row where col C holds a real product name
      abr_row            — product_name_row + 1  (abbreviations, kept for reference only)
      material_start_row — first row in col A with a real ingredient name
    Works regardless of how many header/blank rows appear before the data.
    """
    product_name_row = None
    for r in range(min(10, df.shape[0])):
        v = str(df.iloc[r, 2]).strip()
        if v in ('nan', 'NaN', ''):
            continue
        try:
            float(v); continue
        except ValueError:
            pass
        if v.lower() in _SM_HEADER_WORDS:
            continue
        product_name_row = r
        break
    if product_name_row is None:
        raise ValueError(
            "Standard mix: cannot detect product name row. "
            "Expected a product name in column C within the first 10 rows."
        )
    abr_row = product_name_row + 1
    material_start_row = None
    for r in range(abr_row + 1, min(abr_row + 20, df.shape[0])):
        v = str(df.iloc[r, 0]).strip()
        if v in ('nan', 'NaN', ''):
            continue
        try:
            float(v); continue
        except ValueError:
            pass
        if v.lower() in _SM_HEADER_WORDS:
            continue
        material_start_row = r
        break
    if material_start_row is None:
        raise ValueError(
            "Standard mix: cannot detect material start row. "
            "Expected an ingredient name in column A within 20 rows of the abbreviation row."
        )
    return product_name_row, abr_row, material_start_row


def _parse_standard_mix(sm_bytes):
    """
    Parses any company's standard mix dynamically.

    Matching key = NORMALISED PRODUCT NAME (uppercase, stripped).
    ABRs are intentionally NOT used as keys because:
      - The same ABR can appear for different products (e.g. 'BB' for both
        'BANS BIG' and 'BUFFET BREAD' in some files).
      - ABR spacing is inconsistent across files ('C S' vs 'CS').
      - Product names are unambiguous and stable.

    Returns:
      products  : {norm_product_name: {'name': str, 'ratios': {mat_name: float}}}
      materials : [(mat_name, unit)]
    """
    df = pd.read_excel(BytesIO(sm_bytes), header=None)
    product_name_row, abr_row, material_start_row = _detect_sm_structure(df)

    # Raw materials — col 0 = name, col 1 = unit, from material_start_row downward
    materials = []
    for r in range(material_start_row, df.shape[0]):
        mat_name = str(df.iloc[r, 0]).strip()
        mat_unit = str(df.iloc[r, 1]).strip()
        if mat_name in ('nan', 'NaN', ''):
            break
        try:
            float(mat_name); break
        except ValueError:
            pass
        if mat_name.lower() in _SM_HEADER_WORDS:
            continue
        unit_clean = mat_unit if mat_unit not in ('nan', 'NaN', '') else ''
        materials.append((mat_name, unit_clean))

    if not materials:
        raise ValueError(
            "Standard mix: no raw materials found. "
            "Check that ingredient names start in column A."
        )

    # Products — every 5 columns starting at col 2
    # Keyed by normalised product NAME to avoid all ABR ambiguity
    products = {}
    for c in range(2, df.shape[1], 5):
        name_str = str(df.iloc[product_name_row, c]).strip()
        # Skip empty, numeric, or header-keyword cells
        if name_str in ('nan', 'NaN', ''):
            continue
        try:
            float(name_str); continue
        except ValueError:
            pass
        if name_str.lower() in _SM_HEADER_WORDS:
            continue

        ratios = {}
        for i, (mat_name, _) in enumerate(materials):
            row_idx = material_start_row + i
            if row_idx >= df.shape[0]:
                ratios[mat_name] = 0.0
                continue
            val = df.iloc[row_idx, c]
            try:
                ratios[mat_name] = (
                    float(val) if str(val).strip() not in ('nan', 'NaN', '') else 0.0
                )
            except (ValueError, TypeError):
                ratios[mat_name] = 0.0

        norm = _norm_name(name_str)
        # If two columns share the same normalised name, last one wins (edge case)
        products[norm] = {'name': name_str, 'ratios': ratios}

    if not products:
        raise ValueError(
            "Standard mix: no valid products found. "
            "Check that product names appear in the first detected row."
        )
    return products, materials


def _parse_finished_movement(fm_bytes):
    """
    Reads col B (product name) and col F (EXPECTED) for every row.
    expected_map  keyed by (date, NORM_PRODUCT_NAME)
    fm_products   set of original product name strings (for display in errors)
    """
    wb = openpyxl.load_workbook(BytesIO(fm_bytes), data_only=True)
    ws = wb.active
    expected_map = {}
    fm_products  = set()
    current_date = None

    for r in range(1, ws.max_row + 1):
        date_v = ws.cell(r, 1).value
        name_v = ws.cell(r, 2).value   # col B — product full name
        exp_v  = ws.cell(r, 6).value   # col F — EXPECTED

        if isinstance(date_v, datetime):
            current_date = date_v.date()

        if not name_v or not str(name_v).strip():
            continue
        name_str = str(name_v).strip()

        # Skip obvious header rows
        if name_str.upper() in ('PRODUCTS', 'DETAILS', 'DATE', 'DESCRIPTION'):
            continue

        fm_products.add(name_str)
        if current_date is None:
            continue

        try:
            qty = float(exp_v) if exp_v not in (None, '') else 0.0
        except (ValueError, TypeError):
            qty = 0.0

        key = (current_date, _norm_name(name_str))
        expected_map[key] = expected_map.get(key, 0.0) + qty

    return expected_map, fm_products


def _match_product_name(fm_name, sm_norm_names):
    """
    Match a single finished-movement product name to a standard-mix product name.
    Returns (matched_norm_name_or_None, tag_string).
    Uses: exact → fuzzy (best match, cutoff 0.60).
    No substring check — fuzzy naturally picks the highest-ratio match.
    """
    t = _norm_name(fm_name)
    if t in sm_norm_names:
        return t, 'EXACT'
    hits = difflib.get_close_matches(t, sm_norm_names, n=1, cutoff=0.60)
    if hits:
        return hits[0], 'FUZZY'
    return None, 'NO MATCH'


def _parse_rm_template(rm_bytes):
    """
    Reads the user's blank raw material template.
    Detects block structure (materials per day) dynamically — works for any
    number of materials (1 product kombucha up to 50+ product bakery).
    """
    wb = openpyxl.load_workbook(BytesIO(rm_bytes))
    ws = wb.active

    # First date row = start of day-1 block
    first_data_row = None
    for r in range(1, ws.max_row + 1):
        if isinstance(ws.cell(r, 1).value, datetime):
            first_data_row = r
            break
    if first_data_row is None:
        raise ValueError("Raw material template: no date found in column A.")

    # Second date row → block_size
    second_date_row = None
    for r in range(first_data_row + 1, ws.max_row + 1):
        if isinstance(ws.cell(r, 1).value, datetime):
            second_date_row = r
            break

    if second_date_row:
        block_size = second_date_row - first_data_row
    else:
        # Single-day template: count until first blank in col B
        block_size = 0
        for r in range(first_data_row, ws.max_row + 1):
            if ws.cell(r, 2).value is None and r > first_data_row:
                block_size = r - first_data_row + 1
                break
        if block_size == 0:
            block_size = ws.max_row - first_data_row + 1

    n_materials = block_size - 1   # last row in every block is the blank separator

    # Read material names + units from day-1 block (col B = name, col C = unit)
    tpl_materials = []
    for i in range(n_materials):
        r   = first_data_row + i
        mat = ws.cell(r, 2).value
        unt = ws.cell(r, 3).value
        if mat and str(mat).strip():
            tpl_materials.append((
                i,
                str(mat).strip(),
                str(unt).strip() if unt else ''
            ))

    # Map every date in the template to its block-start row
    date_to_block_start = {}
    for r in range(first_data_row, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, datetime):
            date_to_block_start[v.date()] = r

    return wb, ws, n_materials, tpl_materials, date_to_block_start


def _smart_match_materials(tpl_materials, sm_materials):
    """
    Match raw material names in the template → raw material names in the
    standard mix.  Uses exact then fuzzy (cutoff 0.65).
    Returns match_map {tpl_name: sm_name_or_None}, unmatched list.
    """
    sm_names  = [m[0] for m in sm_materials]
    sm_upper  = [m.upper().strip() for m in sm_names]
    match_map = {}
    unmatched = []

    for _, tpl_name, _ in tpl_materials:
        t = tpl_name.upper().strip()
        matched = None

        # 1. Exact
        if t in sm_upper:
            matched = sm_names[sm_upper.index(t)]

        # 2. Fuzzy — returns best match; no substring shortcut to avoid wrong greedy hits
        if not matched:
            hits = difflib.get_close_matches(t, sm_upper, n=1, cutoff=0.65)
            if hits:
                matched = sm_names[sm_upper.index(hits[0])]

        match_map[tpl_name] = matched
        if not matched:
            unmatched.append(tpl_name)

    return match_map, unmatched


def _calculate_and_fill(wb, ws, tpl_materials, date_to_block_start,
                         match_map, products, expected_map, fm_name_to_sm_norm):
    """
    For each day × each raw material:
      col 9 = Σ ( ratio(material, product) × EXPECTED(product, day) )
              summed over every finished product.

    products      keyed by norm_product_name
    expected_map  keyed by (date, norm_product_name_from_fm)
    fm_name_to_sm_norm maps norm FM product name → norm SM product name
                        (handles fuzzy name differences between files)
    Only col 9 is written; all other formulas in the template are preserved.
    """
    filled = 0
    for date_obj, block_start in sorted(date_to_block_start.items()):
        for offset, tpl_name, _ in tpl_materials:
            sm_mat_name = match_map.get(tpl_name)
            if sm_mat_name is None:
                continue

            total = 0.0
            for norm_fm, exp in expected_map.items():
                if norm_fm[0] != date_obj or exp == 0.0:
                    continue
                # Translate FM product name key → SM product name key
                sm_norm = fm_name_to_sm_norm.get(norm_fm[1])
                if sm_norm is None:
                    continue
                prod = products.get(sm_norm)
                if prod is None:
                    continue
                ratio  = prod['ratios'].get(sm_mat_name, 0.0)
                total += ratio * exp

            if total > 0:
                ws.cell(block_start + offset, 9).value = round(total, 4)
                filled += 1

    return wb, filled


def process_raw_material_movement(sm_file, fm_file, rm_template_file, output_name):
    """Master function called from Streamlit UI."""
    log = []
    def L(msg): log.append(msg)

    try:
        # ── Step 1: Standard Mix ─────────────────────────────────────────────
        L("📋 Step 1 — Parsing standard mix...")
        sm_bytes  = sm_file.read()
        products, sm_materials = _parse_standard_mix(sm_bytes)
        L(f"  ✅ {len(products)} finished product(s) | {len(sm_materials)} raw material(s)")
        L("  📦 Products in standard mix:")
        for norm, p in products.items():
            L(f"       • {p['name']}")

        # ── Step 2: Finished Movement ────────────────────────────────────────
        L("📋 Step 2 — Parsing finished movement (EXPECTED column)...")
        fm_bytes  = fm_file.read()
        expected_map, fm_products = _parse_finished_movement(fm_bytes)
        non_zero  = sum(1 for v in expected_map.values() if v > 0)
        L(f"  ✅ {len(fm_products)} product name(s) | "
          f"{len(expected_map)} date-product pairs ({non_zero} with data)")

        # ── Step 3: Match FM product names → SM product names ────────────────
        L("🔍 Step 3 — Matching finished movement products to standard mix...")
        sm_norm_list = list(products.keys())   # normalised SM product names
        fm_name_to_sm_norm = {}
        missing = []
        fuzzy_warnings = []

        for fm_name in sorted(fm_products):
            matched_norm, tag = _match_product_name(fm_name, sm_norm_list)
            if matched_norm:
                fm_name_to_sm_norm[_norm_name(fm_name)] = matched_norm
                sm_display = products[matched_norm]['name']
                L(f"       ✓ [{tag:8}]  '{fm_name}'  →  '{sm_display}'")
                if tag == 'FUZZY':
                    fuzzy_warnings.append((fm_name, sm_display))
            else:
                missing.append(fm_name)
                L(f"       ✗ [NO MATCH]  '{fm_name}'  →  NOT FOUND in standard mix")

        if missing:
            L(f"  ❌ {len(missing)} product(s) from the finished movement could not be "
              f"matched to any product in the standard mix:")
            for m in missing:
                L(f"       • '{m}'")
            L("  ⚠️  Add these products to the standard mix before processing.")
            return None, log

        if fuzzy_warnings:
            L(f"  ⚠️  {len(fuzzy_warnings)} product(s) matched by fuzzy name — "
              f"please verify these are correct:")
            for fm_n, sm_n in fuzzy_warnings:
                L(f"       '{fm_n}'  →  '{sm_n}'")

        L(f"  ✅ All {len(fm_products)} product(s) matched successfully")

        # ── Step 4: Raw Material Template ────────────────────────────────────
        L("📋 Step 4 — Reading raw material template structure...")
        rm_bytes  = rm_template_file.read()
        wb, ws, n_materials, tpl_materials, date_to_block_start = _parse_rm_template(rm_bytes)
        L(f"  ✅ {len(date_to_block_start)} day(s) | {n_materials} material row(s) per day")
        L("  📄 Template materials:")
        for _, name, unit in tpl_materials:
            L(f"       • {name} ({unit})")

        # ── Step 5: Match template material names → SM material names ────────
        L("🔗 Step 5 — Matching template materials to standard mix...")
        match_map, unmatched = _smart_match_materials(tpl_materials, sm_materials)
        matched_n = sum(1 for v in match_map.values() if v)
        L(f"  ✅ {matched_n}/{len(tpl_materials)} raw material(s) matched")
        for tpl_n, sm_n in match_map.items():
            if sm_n:
                tag = "EXACT" if tpl_n.upper().strip() == sm_n.upper().strip() else "FUZZY"
                L(f"       ✓ [{tag}]  '{tpl_n}'  →  '{sm_n}'")
            else:
                L(f"       ✗ [NO MATCH]  '{tpl_n}'  →  will be left blank in output")
        if unmatched:
            L(f"  ⚠️  {len(unmatched)} raw material(s) unmatched — "
              f"verify spelling between template and standard mix")

        # ── Step 6: Calculate and fill ───────────────────────────────────────
        L("⚙️  Step 6 — Calculating STOCK ISSUED TO PRODUCTION...")
        wb_out, filled = _calculate_and_fill(
            wb, ws, tpl_materials, date_to_block_start,
            match_map, products, expected_map, fm_name_to_sm_norm
        )
        L(f"  ✅ {filled} cell(s) written to STOCK ISSUED TO PRODUCTION (col I)")

        out = BytesIO()
        wb_out.save(out)
        out.seek(0)
        L(f"✅ Complete — ready to download: {output_name}")
        return out, log

    except Exception as e:
        import traceback
        L(f"❌ Error: {str(e)}")
        L(traceback.format_exc())
        return None, log


# ─────────────────────────────────────────────────────────────────────────────
# NAVIGATION
# ─────────────────────────────────────────────────────────────────────────────

st.sidebar.title("Navigation")
tool = st.sidebar.selectbox("Select Tool", [
    "Excel Stock Movement Filler",
    "EFRIS Invoice Enricher",
    "Raw Material Movement Filler",
    "Audit Compliance Checker (Coming Soon)",
    "Financial Report Generator (Coming Soon)",
    "Sales Dashboard (Coming Soon)",
])

st.title("Automation Hub")
st.markdown("Your professional platform for automating tasks.")


# ── TOOL 1 UI ────────────────────────────────────────────────────────────────
if tool == "Excel Stock Movement Filler":
    st.header("Excel Stock Movement Filler")
    output_name   = st.text_input("Output Filename", value="filled_template")
    output_name   = output_name.removesuffix(".xlsx").strip() + ".xlsx"
    template_file = st.file_uploader("Upload Template (.xlsx)", type="xlsx")
    report_file   = st.file_uploader("Upload Movement Report (.xlsx)", type="xlsx")
    damages_file  = st.file_uploader("Upload Damages (.xlsx)", type="xlsx")
    if st.button("Process Files"):
        if template_file and report_file and damages_file:
            with st.spinner("Processing..."):
                out = process_excel(template_file, report_file, damages_file, output_name)
                if out:
                    st.success("Done!")
                    st.download_button("Download", data=out, file_name=output_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.warning("Upload all 3 files.")


# ── TOOL 2 UI ────────────────────────────────────────────────────────────────
elif tool == "EFRIS Invoice Enricher":
    st.header("EFRIS Invoice Enricher")
    st.markdown("""
    Upload your **Purchases Report** (.xlsx) with columns **FDN** and **Description of Goods**.
    The tool opens each invoice on EFRIS, reads the PDF, and fills in
    **Quantity**, **Unit Measure**, and **Unit Price** for every row.
    > Duplicate FDNs are only fetched once.
    """)
    col1, col2 = st.columns([2, 1])
    with col1:
        purchases_file = st.file_uploader("Upload Purchases Report (.xlsx)", type=["xlsx"], key="ef_up")
    with col2:
        out_name = st.text_input("Output Filename", value="enriched_purchases", key="ef_out")
        out_name = out_name.removesuffix(".xlsx").strip() + ".xlsx"
    if purchases_file:
        try:
            prev = pd.read_excel(purchases_file, nrows=5)
            purchases_file.seek(0)
            st.markdown("**Preview:**")
            st.dataframe(prev, use_container_width=True)
            missing = {"FDN", "Description of Goods"} - set(prev.columns)
            if missing:
                st.error(f"Missing columns: {missing}")
                purchases_file = None
        except Exception as e:
            st.error(str(e)); purchases_file = None
    if st.button("🚀 Start Enrichment", disabled=(purchases_file is None), key="ef_run"):
        st.markdown("---")
        prog   = st.progress(0, text="Starting...")
        log_ph = st.empty()
        try:
            df       = pd.read_excel(purchases_file)
            enriched = run_efris_enrichment(df, log_ph, prog)
            prog.progress(1.0, text="✅ Done!")
            st.success("Complete!")
            st.download_button("⬇️ Download Enriched Excel",
                data=build_output_excel(enriched), file_name=out_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            filled = enriched["Quantity"].notna().sum()
            st.info(f"📊 {filled} / {len(enriched)} rows enriched.")
        except Exception as e:
            st.error(f"Error: {e}")


# ── TOOL 3 UI ────────────────────────────────────────────────────────────────
elif tool == "Raw Material Movement Filler":
    st.header("Raw Material Movement Filler")
    st.markdown("""
    Calculates **STOCK ISSUED TO PRODUCTION** for every raw material, every day —
    derived from the **EXPECTED** quantities in the finished goods movement and your
    company's **standard mix** ratios.

    **How it works:**
    - For each day and each raw material:
      **Issued = Σ (input/unit ratio × EXPECTED quantity)** across all finished products
    - The result is written into the *STOCK ISSUED TO PRODUCTION* column of your
      raw material template. All existing formulas (Bal b/f carry-forwards, totals, etc.)
      are preserved.

    > ⚠️  All products in the finished movement **must** be present in the standard mix.
    > The tool will list any missing ones and stop before making changes.
    """)

    st.markdown("---")
    c1, c2, c3 = st.columns(3)
    with c1:
        sm_file = st.file_uploader(
            "① Standard Mix (.xlsx)",
            type="xlsx", key="rm_sm",
            help="Your company's standard mix with products and raw material ratios."
        )
    with c2:
        fm_file = st.file_uploader(
            "② Finished Goods Movement (.xlsx)",
            type="xlsx", key="rm_fm",
            help="Output of Tool 1 — must contain the EXPECTED column (col F)."
        )
    with c3:
        rm_file = st.file_uploader(
            "③ Raw Material Template (.xlsx)",
            type="xlsx", key="rm_tpl",
            help="Your blank raw material movement template (company-specific)."
        )

    out_name_rm = st.text_input("Output Filename", value="raw_material_filled", key="rm_out")
    out_name_rm = out_name_rm.removesuffix(".xlsx").strip() + ".xlsx"

    # Show previews of each uploaded file
    if sm_file:
        try:
            sm_prev = pd.read_excel(sm_file, header=None, nrows=4)
            sm_file.seek(0)
            with st.expander("Preview: Standard Mix (first 4 rows)"):
                st.dataframe(sm_prev, use_container_width=True)
        except Exception:
            pass

    if fm_file:
        try:
            fm_prev = pd.read_excel(fm_file, header=None, nrows=8)
            fm_file.seek(0)
            with st.expander("Preview: Finished Goods Movement (first 8 rows)"):
                st.dataframe(fm_prev, use_container_width=True)
        except Exception:
            pass

    if rm_file:
        try:
            rm_prev = pd.read_excel(rm_file, header=None, nrows=8)
            rm_file.seek(0)
            with st.expander("Preview: Raw Material Template (first 8 rows)"):
                st.dataframe(rm_prev, use_container_width=True)
        except Exception:
            pass

    all_uploaded = sm_file and fm_file and rm_file
    if not all_uploaded:
        st.info("Upload all three files above to proceed.")

    if st.button("⚙️  Generate Raw Material Movement", disabled=not all_uploaded, key="rm_run"):
        st.markdown("---")
        log_ph = st.empty()
        with st.spinner("Processing..."):
            out_bytes, log_lines = process_raw_material_movement(
                sm_file, fm_file, rm_file, out_name_rm
            )
        # Show log
        log_ph.markdown(
            '<div class="log-box">' + "<br>".join(log_lines) + "</div>",
            unsafe_allow_html=True
        )
        if out_bytes:
            st.success("✅ Processing complete!")
            st.download_button(
                "⬇️ Download Filled Raw Material Movement",
                data=out_bytes,
                file_name=out_name_rm,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.error("Processing failed. See the log above for details.")


# ── COMING SOON ──────────────────────────────────────────────────────────────
elif "Coming Soon" in tool:
    st.info("Feature coming soon.")

st.sidebar.markdown("---")
st.sidebar.info("Powered by Streamlit on Railway.")
